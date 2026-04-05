# --- IMPORTAÇÕES ---

import os
import shutil
import pandas as pd
from time import sleep
from datetime import datetime
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException


# --- CONFIGURAÇÕES ---
# Em um projeto real, use variáveis de ambiente ou um arquivo .env
USER_MES = "fredsoncabral"
PASS_MES = "123xfc"
URL_LOGIN = "http://xfc/pcfui#/page/Login"
URL_REPORTE = "http://xfc/pcfui#/page/EMIViewer?reportCode=xfc"

PATHS = {
    "base_local": "/projetos/data/base.xlsx",
    "base_rede": "/indicadores/prod/xfc/base.xlsx",
    "download_temp": "/temp/downloads/report.xlsx",
    "projeto_data": "/projetos/data/report.xlsx",
    "processado": "/projetos/data/xfc_processado.xlsx"
}


# --- AUTOMAÇÃO ETL ---

class XfcAutomation:
    def __init__(self):
        options = webdriver.EdgeOptions()
        self.driver = webdriver.ChromiumEdge(options=options)
        self.wait = WebDriverWait(self.driver, 20)

    def login(self):
        print("[*] Iniciando Login...")
        self.driver.get(URL_LOGIN)
        self.driver.maximize_window()
        
        self.wait.until(EC.element_to_be_clickable((By.ID, "username"))).send_keys(USER_MES)
        self.wait.until(EC.element_to_be_clickable((By.ID, "txtPassword"))).send_keys(PASS_MES)
        self.driver.find_element(By.ID, "btnLogin").click()
        
        # Lidar com modais de aviso
        try:
            self.wait.until(EC.element_to_be_clickable((By.ID, "bModalClose"))).click()
            self.wait.until(EC.element_to_be_clickable((By.ID, "bModalOk"))).click()
        except TimeoutException:
            pass

    def extrair_dados(self, data_inicio, data_fim):
        print(f"[*] Extraindo dados de {data_inicio} até {data_fim}...")
        self.driver.get(URL_REPORTE)
        
        self.wait.until(EC.element_to_be_clickable((By.ID, "bParams"))).click()
        
        # Preencher Datas
        self.wait.until(EC.presence_of_element_located((By.XPATH, "//div[@pname='dtInicial']//input"))).send_keys(data_inicio)
        self.wait.until(EC.presence_of_element_located((By.XPATH, "//div[@pname='dtFinal']//input"))).send_keys(data_fim)
        
        # Lógica do Dropdown "Sim"
        dropdowns = self.wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, ".param-value .dropdown")))
        dropdowns[1].find_element(By.CLASS_NAME, "dropdown-toggle").click()
        
        opcoes = self.wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "ul.dropdown-menu li")))
        for opcao in opcoes:
            if "Sim" in opcao.text:
                opcao.click()
                break
        
        self.driver.find_element(By.ID, "refresh-button").click()
        print("[!] Aguardando processamento do relatório (25s)...")
        sleep(25)
        
        # Exportar
        self.wait.until(EC.element_to_be_clickable((By.XPATH, "//div[@class='dropdown']/a[@class='dropdown-toggle']"))).click()
        self.driver.find_element(By.ID, "bExpandExportExcel").click()
        sleep(5)
        self.driver.quit()

def tratar_dados():
    print("[*] Tratando dados com Pandas...")
    # Limpeza inicial do Report
    df = pd.read_excel(PATHS["projeto_data"], header=None)
    df = df.iloc[10:, 3:].reset_index(drop=True)
    df = df.dropna(axis=1, how='all')
    
    # Seleção de colunas específicas (ajuste os índices se necessário)
    df = df.iloc[:, [0, 8, 3, 9, 10, 19, 11]] 
    
    df.to_excel(PATHS["processado"], index=False, header=False)

def atualizar_base_excel():
    print("[*] Atualizando Base de Dados Mestra...")
    df_novo = pd.read_excel(PATHS["processado"], header=None).dropna(how='all')
    
    wb = load_workbook(PATHS["base_local"])
    ws = wb["F_Base_de_Dados"]
    
    prox_linha = ws.max_row + 1
    
    for r_idx, row in enumerate(df_novo.values, start=prox_linha):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    wb.save(PATHS["base_local"])

def main():
    print("=== SISTEMA DE AUTOMAÇÃO CALLIDUS ===")
    data_i = input("Data Inicial (DDMMYYYY): ")
    data_f = input("Data Final (DDMMYYYY): ")

    try:
        # 1. Preparação
        if os.path.exists(PATHS["base_rede"]):
            shutil.copy2(PATHS["base_rede"], PATHS["base_local"])
        
        # 2. Automação Web
        bot = XfcAutomation()
        bot.login()
        bot.extrair_dados(data_i, data_f)
        
        # 3. Mover arquivo baixado
        if os.path.exists(PATHS["download_temp"]):
            shutil.move(PATHS["download_temp"], PATHS["projeto_data"])
        
        # 4. Processamento
        tratar_dados()
        atualizar_base_excel()
        
        # 5. Sincronização Final
        shutil.copy2(PATHS["base_local"], PATHS["base_rede"])
        print("\n[OK] Processo finalizado com sucesso!")
        
    except Exception as e:
        print(f"\n[ERRO] Ocorreu uma falha: {e}")

if __name__ == "__main__":
    main()